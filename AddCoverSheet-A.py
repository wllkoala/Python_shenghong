import os
import traceback
from datetime import datetime
from sys import exit
from tkinter import Tk, filedialog, messagebox

from openpyxl import load_workbook
from PyPDF4 import PdfFileReader, PdfFileWriter
from win32com.client import DispatchEx
from xlrd import open_workbook

root = Tk()
root.withdraw()


class AddCoverSheet():
    name_list = []
    doc_code = []
    doc_rev = []
    final_names = []
    file_lists = []

    def __init__(self, file_dir):
        self.file_dir = file_dir

    def start_to_run(self):
        '''开始运行'''
        self.create_folder()
        self.get_name_list()
        self.cover_sheet()
        self.conversion()
        self.merge_doc()
        messagebox.showinfo("Complete!", "全部文件已完成！")

    def create_folder(self):
        '''创建所需文件夹'''
        if not os.path.exists("input"):
            os.mkdir("input")
        if not os.path.exists("output"):
            os.mkdir("output")
        else:
            self.del_file("output")
        if not os.path.exists("tmp"):
            os.mkdir("tmp")
        else:
            self.del_file("tmp")

    def del_file(self, file_dir):
        '''删除已有文件'''
        for file in os.listdir(file_dir):
            path_file = os.path.join(file_dir, file)
            if os.path.isfile(path_file):
                os.remove(path_file)

    def get_name_list(self):
        '''获取待添加封面文件目录'''
        for name in os.listdir("input"):
            if name.endswith(".pdf"):
                if name.find('_') > 0:
                    self.doc_code.append(name.split("_")[0])
                    self.doc_rev.append(name.split("_")[1][1:3])
                    self.name_list.append(name)
                    print("需要添加封面的文件：", len(self.name_list),
                          name.split(".")[0])
        print("=><=" * 25)

    def write_data(self, data, rev):
        '''向excel中写入数据'''
        # 如果是图纸文件使用图纸封面，不适则使用文件封面
        rev_fls = ("0", "A", "B", 'C', 'D', 'E', 'F',
                   'G', 'H', "I", "J", "K", "L", "M", "N")
        if data[7].upper() == "Y":
            wb = load_workbook('Shenghong-drawing.xlsx')
            ws = wb.worksheets[0]
            ws.cell(2, 4).value = data[0]
            ws.cell(6, 4).value = data[1]
            ws.cell(8, 4).value = data[2]
            ws.cell(10, 4).value = data[3]
            ws.cell(12, 4).value = data[4] + "_Rev " + rev_fls[int(rev)]
            ws.cell(14, 4).value = data[5] + "_A00"
            ws.cell(16, 4).value = data[6]
            ws.cell(18, 4).value = datetime.now()
        else:
            wb = load_workbook('Shenghong-doc.xlsx')
            ws = wb.worksheets[0]
            ws.cell(2, 4).value = data[0]
            ws.cell(6, 4).value = data[1]
            ws.cell(8, 4).value = data[2]
            ws.cell(10, 4).value = data[3]
            ws.cell(12, 4).value = data[4] + "_Rev " + rev_fls[int(rev)]
            ws.cell(14, 4).value = data[5] + "_A00"
            ws.cell(16, 4).value = data[6]
            for row in range(20, 30):
                if ws.cell(row, 1).value == "版次\nRev.":
                    for i in range(0, 1):
                        ws.cell(row - 1 - i, 1).value = rev_fls[i]
                        ws.cell(row - 1 - i, 2).value = "For Review"
                        ws.cell(row - 1 - i, 3).value = data[8]
                        ws.cell(row - 1 - i, 5).value = datetime.now()
                        ws.cell(row - 1 - i, 6).value = data[9]
                        ws.cell(row - 1 - i, 7).value = datetime.now()
                        ws.cell(row - 1 - i, 8).value = data[10]
                        ws.cell(row - 1 - i, 9).value = datetime.now()
        file_name = os.path.join("tmp", data[4] + ".xlsx")
        wb.save(file_name)
        final_name = data[5] + '_A00' + rev + '_' + data[6]
        return final_name

    def conversion(self):
        '''转换封面EXCEL为PDF'''
        xlApp = DispatchEx("Excel.Application")
        xlApp.Visible = False
        xlApp.DisplayAlerts = 0
        for name_list_index, name_list in enumerate(self.doc_code):
            print('当前文件转换进度',
                  name_list_index + 1, "/", len(self.doc_code))
            exportfile = name_list
            filenames = exportfile.split('.')[0] + '.xlsx'
            filename = filenames.replace("input", "tmp")
            books = xlApp.Workbooks.Open(filename, False)
            books.ExportAsFixedFormat(0, exportfile)
            books.Close(False)
            print('封面转为PDF文件：', exportfile)
        xlApp.Quit()
        print('封面转为PDF文件完成')
        print("=><=" * 25)

    def merge_doc(self):
        '''合并封面和文件'''
        self.file_lists = list(
            zip(self.doc_code, self.name_list, self.final_names))
        for pdfnames in self.file_lists:
            output = PdfFileWriter()
            for pdfname in pdfnames[0:2]:
                input = PdfFileReader(open(pdfname, "rb"), strict=False)
                pageCount = input.getNumPages()
                for iPage in range(0, pageCount):
                    output.addPage(input.getPage(iPage))
            pdfoutname = str(pdfnames[2])
            outputStream = open(pdfoutname, "wb")
            output.write(outputStream)
            outputStream.close()
            print("文件合并完成：", pdfoutname)
        print("文件合并完成！")
        print("=><=" * 25)

    def cover_sheet(self):
        '''生成excel版封面'''
        excel_file = filedialog.askopenfilename(title="选择编辑好的VDL", filetypes=[
                                                ("Excel", "*.xlsx"), ("All files", "*")])

        wb = open_workbook(excel_file)
        ws = wb.sheets()[0]
        n_rows_num = ws.nrows
        print("需要生成封面文件数：", len(self.doc_code))
        for n in range(len(self.doc_code)):
            for i in range(0, n_rows_num):
                data = ws.row_values(i)
                for j in range(0, len(data)):
                    temp_data = str(data[j])
                    if self.doc_code[n] == temp_data:
                        final_name = self.write_data(
                            data=data, rev=self.doc_rev[n])
                        print("当前封面生成进度：", n + 1, "/",
                              len(self.doc_code))
                        print("文件封面已完成", data[4], "_R", self.doc_rev[n])
                        self.final_names.append(
                            os.path.join(self.file_dir, "output", final_name + '.pdf'))
            self.doc_code[n] = os.path.join(
                self.file_dir, "input", self.doc_code[n] + '.pdf')
            self.name_list[n] = os.path.join(
                self.file_dir, "input", self.name_list[n])
        print("=><=" * 25)


if __name__ == "__main__":
    try:
        file_dir = os.getcwd()
        print("=><=" * 25)
        print("当前路径：", file_dir)
        print("=><=" * 25)
        merge = AddCoverSheet(file_dir)
        merge.start_to_run()
    except Exception as err:
        messagebox.showerror("Warning!", err)
        with open("d:/a.log", "a") as f:
            traceback.print_exc(file=f)
        print(err)
        exit()
